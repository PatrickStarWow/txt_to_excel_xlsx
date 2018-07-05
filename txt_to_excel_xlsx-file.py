# encoding:utf-8  

from openpyxl import *  
from optparse import OptionParser
import sys

def txt_to_excel(inputfile):
    # 读取参数路径文件  
    f = open(inputfile, 'r')
    line = f.read()
    # 创建workbook  
    wb = Workbook()  
    # 增加一个sheet页'result'
    ws = wb.create_sheet('result', 0)

    # 以'*'分割，获取每行数据
    arr_line = line[0:].split('\n')
    for i in range(len(arr_line)):  
        # 对行数据进行遍历，获取行数据元素元组  
        arr_cell = arr_line[i].split(' ')  
        for j in range(len(arr_cell)):  
            # 写入数据
            ws.cell(row=i+1, column=j+1, value=arr_cell[j])
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["D"].width = 28

    inputfile_excel=inputfile.replace('txt', 'xlsx')
    wb.save(inputfile_excel)


if __name__== "__main__":

    parser = OptionParser(description='ports&*weak password scanner. company:mogu security. teams:xdsec. author: wilson ')
    parser.add_option("--file", action="store", dest="file", type="string",
                      help='选择要转换的txt文件')
    (options, args) = parser.parse_args()
    file = options.file
    print file
    txt_to_excel(file)